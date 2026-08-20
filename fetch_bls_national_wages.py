"""
Fetches current BLS OEWS *national* median wage (plus 10th/90th percentile
and employment) for an arbitrary list of occupations -- for stats that cite
a spread across several distinct SOC codes rather than one occupation
broken out by state (that's what fetch_bls_wages.py is for).

First use case: the "Can you get into med school with a biomedical
engineering degree?" article's physician/surgeon pay-range sentence
("general practitioners can earn $200,000 or more ... surgeons or
anesthesiologists can earn $300,000+"), which cites several physician
specialties, not one occupation across states.

Usage:
    python fetch_bls_national_wages.py "<SOC>:<Label>,<SOC>:<Label>,..." [output.xlsx] [--baseline=path.json]
    python fetch_bls_national_wages.py "29-1215:Family Medicine Physicians,29-1216:General Internal Medicine Physicians,29-1221:Pediatricians General,29-1249:Surgeons All Other,29-1211:Anesthesiologists" "Physician Wages.xlsx"

--baseline points at a JSON file shaped like:
    {"values": {"Family Medicine Physicians": {"median": 244180, ...}, ...}}

Reuses the series-ID construction, suppressed-data handling, and batching
from fetch_bls_wages.py rather than duplicating it -- see that file for the
BLS API details (key, rate limits, etc).
"""

import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from fetch_bls_wages import DATATYPE, fetch_series_batch, national_series_id, parse_bls_number

REVIEW_THRESHOLD_PCT = 10


def parse_occupations(spec):
    """'29-1215:Family Medicine Physicians,...' -> [(soc, label), ...]"""
    occupations = []
    for entry in spec.split(","):
        entry = entry.strip()
        if not entry:
            continue
        soc, _, label = entry.partition(":")
        occupations.append((soc.strip(), label.strip() or soc.strip()))
    return occupations


def fetch_national_wages(occupations):
    series_ids = []
    for soc, _ in occupations:
        for dt_key in ("employment", "10th_pctile", "median", "90th_pctile"):
            series_ids.append(national_series_id(soc, DATATYPE[dt_key]))

    results_raw = fetch_series_batch(series_ids)

    rows = []
    for soc, label in occupations:
        row = {"soc": soc, "label": label}
        notes = []
        for dt_key in ("employment", "10th_pctile", "median", "90th_pctile"):
            sid = national_series_id(soc, DATATYPE[dt_key])
            if sid in results_raw:
                raw_value, year, footnote = results_raw[sid]
                value = parse_bls_number(raw_value)
                if value is not None:
                    row[dt_key] = value
                    row["year"] = year
                elif footnote:
                    notes.append(f"{dt_key}: {footnote}")
        if notes:
            row["note"] = "; ".join(notes)
        rows.append(row)
    return rows


def load_baseline(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)["values"]


def apply_baseline(rows, baseline_values):
    for row in rows:
        baseline = baseline_values.get(row["label"])
        if not baseline or "median" not in baseline or row.get("median") is None:
            continue
        row["baseline_median"] = baseline["median"]
        pct = (row["median"] - baseline["median"]) / baseline["median"] * 100
        row["pct_change_median"] = round(pct, 1)
        if abs(pct) > REVIEW_THRESHOLD_PCT:
            row["review"] = (f"{pct:+.1f}% vs baseline -- larger than typical "
                              f"annual movement, worth a second look")


def save_to_xlsx(rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "BLS National Wages"
    has_baseline = any("baseline_median" in row for row in rows)

    header = ["Occupation", "SOC Code", "Employment", "10th Percentile",
              "Median", "90th Percentile", "Data Year"]
    if has_baseline:
        header += ["Baseline Median", "% Change", "Review"]
    header.append("Note")

    ws.append(header)
    for col in range(1, len(header) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    for row in rows:
        line = [
            row["label"], row["soc"], row.get("employment"),
            row.get("10th_pctile"), row.get("median"), row.get("90th_pctile"),
            row.get("year"),
        ]
        if has_baseline:
            line += [row.get("baseline_median"), row.get("pct_change_median"), row.get("review")]
        line.append(row.get("note"))
        ws.append(line)

    widths = [32, 12, 12, 16, 14, 16, 12]
    if has_baseline:
        widths += [16, 12, 45]
    widths.append(40)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wb.save(out_path)


def main():
    baseline_path = None
    positional = []
    for arg in sys.argv[1:]:
        if arg.startswith("--baseline="):
            baseline_path = arg.split("=", 1)[1]
        else:
            positional.append(arg)

    if len(positional) < 1:
        print('Usage: python fetch_bls_national_wages.py "<SOC>:<Label>,..." [output.xlsx] [--baseline=path.json]')
        sys.exit(1)

    occupations = parse_occupations(positional[0])
    out_path = positional[1] if len(positional) > 1 else "BLS National Wages.xlsx"

    if not os.environ.get("BLS_API_KEY"):
        print("Note: no BLS_API_KEY set -- unregistered rate limit is 25 requests/day.")

    print(f"Fetching national wage data for {len(occupations)} occupation(s)...")
    rows = fetch_national_wages(occupations)

    if baseline_path:
        apply_baseline(rows, load_baseline(baseline_path))
    save_to_xlsx(rows, out_path)

    def fmt(value):
        return f"${value:,}" if isinstance(value, int) else "?"

    print(f"\nDone. Saved to {out_path}")
    for row in rows:
        line = f"  {row['label']:38s} median={fmt(row.get('median'))}"
        if row.get("pct_change_median") is not None:
            line += f"  [{row['pct_change_median']:+.1f}% vs baseline]"
        if row.get("review"):
            line += f"  ** REVIEW **"
        if row.get("note"):
            line += f"  [{row['note']}]"
        print(line)


if __name__ == "__main__":
    main()
