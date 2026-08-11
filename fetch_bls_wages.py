"""
Fetches current BLS OEWS wage data for a given SOC (occupation) code:
national percentiles plus a fixed, editorially-chosen list of states.

Important: the state list is NOT auto-selected by employment count. Per
Lydia (2026-08-11), the original states were chosen to represent a
spectrum of cost of living, not to be the states with the most workers in
the occupation -- so the state list is always an explicit input here, one
you (or the pillar's methodology doc) decide, not something this script
guesses at.

Usage:
    python fetch_bls_wages.py <SOC-code> <states-comma-separated> [output.xlsx] [--baseline=path.json]
    python fetch_bls_wages.py 17-2031 "California,Massachusetts,Minnesota,Ohio,Texas,Washington" "BME Wage Data.xlsx" --baseline=reference/bme_wage_baseline_2024.json

--baseline points at a JSON file shaped like:
    {"values": {"U.S.": {"10th_pctile": 71860, "median": 106950, "90th_pctile": 165160}, ...}}
When given, the output gets a Baseline Median / % Change (Median) / Review
column -- so you don't have to go dig up prior numbers yourself to sanity
-check a fresh pull. A >10% swing on the median gets flagged for review,
since state-level OEWS estimates for a niche occupation are small-sample
and can move more than you'd expect year to year even when nothing is
actually wrong.

No API key required -- the BLS public API works unregistered, but is
rate-limited to 25 queries/day and 25 series per query. Set the
BLS_API_KEY environment variable (free, instant registration at
https://data.bls.gov/registrationEngine/) to raise the limit to 500
queries/day and 50 series per query.
"""

import json
import os
import sys
import time

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

BLS_API = "https://api.bls.gov/publicAPI/v2/timeseries/data/"
BATCH_SIZE = 25             # unregistered BLS API limit per request
BATCH_SIZE_REGISTERED = 50  # limit per request with a registered key

DATATYPE = {
    "employment": "01",
    "10th_pctile": "11",
    "median": "13",
    "90th_pctile": "15",
}

# name -> (2-digit FIPS, Census Bureau region)
STATE_INFO = {
    "Alabama": ("01", "South"), "Alaska": ("02", "West"),
    "Arizona": ("04", "West"), "Arkansas": ("05", "South"),
    "California": ("06", "West"), "Colorado": ("08", "West"),
    "Connecticut": ("09", "Northeast"), "Delaware": ("10", "South"),
    "District of Columbia": ("11", "South"), "Florida": ("12", "South"),
    "Georgia": ("13", "South"), "Hawaii": ("15", "West"),
    "Idaho": ("16", "West"), "Illinois": ("17", "Midwest"),
    "Indiana": ("18", "Midwest"), "Iowa": ("19", "Midwest"),
    "Kansas": ("20", "Midwest"), "Kentucky": ("21", "South"),
    "Louisiana": ("22", "South"), "Maine": ("23", "Northeast"),
    "Maryland": ("24", "South"), "Massachusetts": ("25", "Northeast"),
    "Michigan": ("26", "Midwest"), "Minnesota": ("27", "Midwest"),
    "Mississippi": ("28", "South"), "Missouri": ("29", "Midwest"),
    "Montana": ("30", "West"), "Nebraska": ("31", "Midwest"),
    "Nevada": ("32", "West"), "New Hampshire": ("33", "Northeast"),
    "New Jersey": ("34", "Northeast"), "New Mexico": ("35", "West"),
    "New York": ("36", "Northeast"), "North Carolina": ("37", "South"),
    "North Dakota": ("38", "Midwest"), "Ohio": ("39", "Midwest"),
    "Oklahoma": ("40", "South"), "Oregon": ("41", "West"),
    "Pennsylvania": ("42", "Northeast"), "Rhode Island": ("44", "Northeast"),
    "South Carolina": ("45", "South"), "South Dakota": ("46", "Midwest"),
    "Tennessee": ("47", "South"), "Texas": ("48", "South"),
    "Utah": ("49", "West"), "Vermont": ("50", "Northeast"),
    "Virginia": ("51", "South"), "Washington": ("53", "West"),
    "West Virginia": ("54", "South"), "Wisconsin": ("55", "Midwest"),
    "Wyoming": ("56", "West"),
}


def series_id(area_type, area_code7, soc_code, datatype_code):
    occ = soc_code.replace("-", "")
    return f"OEU{area_type}{area_code7}000000{occ}{datatype_code}"


def national_series_id(soc_code, datatype_code):
    return series_id("N", "0000000", soc_code, datatype_code)


def state_series_id(state_name, soc_code, datatype_code):
    fips, _ = STATE_INFO[state_name]
    return series_id("S", fips + "00000", soc_code, datatype_code)


def parse_bls_number(raw):
    """BLS uses placeholders like "-" for suppressed/unavailable data
    (e.g. a state with too few employees in an occupation to report)."""
    cleaned = raw.replace(",", "").strip()
    return int(cleaned) if cleaned.isdigit() else None


def fetch_series_batch(series_ids):
    """Returns {series_id: (value, year)} for the latest data point of
    each series. Series with no data are omitted."""
    results = {}
    batch_size = BATCH_SIZE_REGISTERED if os.environ.get("BLS_API_KEY") else BATCH_SIZE
    for i in range(0, len(series_ids), batch_size):
        chunk = series_ids[i:i + batch_size]
        body = {"seriesid": chunk}
        if os.environ.get("BLS_API_KEY"):
            body["registrationkey"] = os.environ["BLS_API_KEY"]
        resp = requests.post(BLS_API, json=body, timeout=30)
        resp.raise_for_status()
        payload = resp.json()
        if payload.get("status") != "REQUEST_SUCCEEDED":
            raise RuntimeError(f"BLS API error: {payload.get('message')}")
        for series in payload["Results"]["series"]:
            data_points = series.get("data", [])
            if data_points:
                latest = data_points[0]  # BLS returns most recent first
                footnote = next(
                    (f["text"] for f in latest.get("footnotes", []) if f.get("text")), None)
                results[series["seriesID"]] = (latest["value"], latest["year"], footnote)
        time.sleep(0.5)
    return results


def fetch_wage_data(soc_code, states):
    for name in states:
        if name not in STATE_INFO:
            raise ValueError(f"Unknown state name: {name!r}")

    employment_series = [national_series_id(soc_code, DATATYPE["employment"])]
    for name in states:
        employment_series.append(state_series_id(name, soc_code, DATATYPE["employment"]))

    print(f"Fetching employment counts for {len(states)} state(s)...")
    employment_results = fetch_series_batch(employment_series)

    employment_by_state = {}
    for name in states:
        sid = state_series_id(name, soc_code, DATATYPE["employment"])
        if sid in employment_results:
            value = parse_bls_number(employment_results[sid][0])
            if value is not None:
                employment_by_state[name] = value

    national_employment_sid = national_series_id(soc_code, DATATYPE["employment"])
    national_employment = None
    if national_employment_sid in employment_results:
        national_employment = parse_bls_number(employment_results[national_employment_sid][0])


    wage_series = []
    for dt_code in (DATATYPE["10th_pctile"], DATATYPE["median"], DATATYPE["90th_pctile"]):
        wage_series.append(national_series_id(soc_code, dt_code))
    for name in states:
        for dt_code in (DATATYPE["10th_pctile"], DATATYPE["median"], DATATYPE["90th_pctile"]):
            wage_series.append(state_series_id(name, soc_code, dt_code))

    print("Fetching wage percentiles...")
    wage_results = fetch_series_batch(wage_series)

    def get_wage(area_series_id_fn, name_or_none):
        row = {}
        notes = []
        for label in ("10th_pctile", "median", "90th_pctile"):
            sid = area_series_id_fn(soc_code, DATATYPE[label])
            if sid in wage_results:
                raw_value, year, footnote = wage_results[sid]
                value = parse_bls_number(raw_value)
                if value is not None:
                    row[label] = value
                    row["year"] = year
                elif footnote:
                    notes.append(f"{label}: {footnote}")
        if notes:
            row["note"] = "; ".join(notes)
        return row

    rows = []
    national_row = get_wage(national_series_id, None)
    national_row["geography"] = "U.S."
    national_row["region"] = "National"
    national_row["employment"] = national_employment
    rows.append(national_row)

    for name in states:
        row = get_wage(lambda soc, dt, n=name: state_series_id(n, soc, dt), name)
        row["geography"] = name
        row["region"] = STATE_INFO[name][1]
        row["employment"] = employment_by_state.get(name)
        rows.append(row)

    return rows


REVIEW_THRESHOLD_PCT = 10


def load_baseline(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)["values"]


def apply_baseline(rows, baseline_values):
    """Annotates each row with its baseline median, % change, and a review
    flag for swings past REVIEW_THRESHOLD_PCT -- state-level OEWS estimates
    for a niche occupation are small-sample and can move more than you'd
    expect year to year even when nothing is actually wrong, so a big swing
    is worth a second look, not an automatic "this must be broken"."""
    for row in rows:
        baseline = baseline_values.get(row.get("geography"))
        if not baseline or "median" not in baseline or row.get("median") is None:
            continue
        row["baseline_median"] = baseline["median"]
        pct = (row["median"] - baseline["median"]) / baseline["median"] * 100
        row["pct_change_median"] = round(pct, 1)
        if abs(pct) > REVIEW_THRESHOLD_PCT:
            row["review"] = (f"{pct:+.1f}% vs baseline -- larger than typical "
                              f"annual movement, worth a second look")


def save_to_xlsx(rows, soc_code, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "BLS Wage Data"
    has_baseline = any("baseline_median" in row for row in rows)

    header = ["Geography", "Region", "Employment", "10th Percentile",
              "Median", "90th Percentile", "Data Year"]
    if has_baseline:
        header += ["Baseline Median", "% Change (Median)", "Review"]
    header.append("Note")

    ws.append([f"SOC {soc_code}"])
    ws.append(header)
    for col in range(1, len(header) + 1):
        ws.cell(row=2, column=col).font = Font(bold=True)

    for row in rows:
        line = [
            row.get("geography"), row.get("region"), row.get("employment"),
            row.get("10th_pctile"), row.get("median"), row.get("90th_pctile"),
            row.get("year"),
        ]
        if has_baseline:
            line += [row.get("baseline_median"), row.get("pct_change_median"), row.get("review")]
        line.append(row.get("note"))
        ws.append(line)

    widths = [22, 12, 12, 16, 14, 16, 12]
    if has_baseline:
        widths += [16, 18, 45]
    widths.append(40)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)


def main():
    baseline_path = None
    positional = []
    for arg in sys.argv[1:]:
        if arg.startswith("--baseline="):
            baseline_path = arg.split("=", 1)[1]
        else:
            positional.append(arg)

    if len(positional) < 2:
        print('Usage: python fetch_bls_wages.py <SOC-code> "State One,State Two,..." '
              '[output.xlsx] [--baseline=path.json]')
        sys.exit(1)

    soc_code = positional[0]
    states = [s.strip() for s in positional[1].split(",") if s.strip()]
    out_path = positional[2] if len(positional) > 2 else f"BLS Wage Data - {soc_code}.xlsx"

    rows = fetch_wage_data(soc_code, states)
    if baseline_path:
        apply_baseline(rows, load_baseline(baseline_path))
    save_to_xlsx(rows, soc_code, out_path)

    def fmt(value):
        return f"${value:,}" if isinstance(value, int) else "?"

    print(f"\nDone. Saved {len(rows)} row(s) to {out_path}")
    for row in rows:
        line = (f"  {row['geography']:20s} median={fmt(row.get('median'))} "
                f"10th={fmt(row.get('10th_pctile'))} 90th={fmt(row.get('90th_pctile'))} "
                f"(employment: {row.get('employment', '?')})")
        if row.get("pct_change_median") is not None:
            line += f"  [{row['pct_change_median']:+.1f}% vs baseline]"
        if row.get("review"):
            line += f"  ** REVIEW: {row['review']} **"
        if row.get("note"):
            line += f"  [{row['note']}]"
        print(line)


if __name__ == "__main__":
    main()
