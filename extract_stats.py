"""
Pillar article stat extractor.

Scans public pillar-cluster article pages and flags statistics likely to
need an annual refresh: numbers tied to an explicit year (e.g. "$106,950
in 2024") as well as numbers attached to a stats keyword even with no
year in the same sentence (e.g. "median pay is $106,950").

Usage:
    python extract_stats.py <urls.txt | urls_dir> [output.xlsx]

Output defaults to "Pillar Stat Review - <Mon DD YYYY>.xlsx" (today's date)
in the current directory if no output path is given.

If given a single file: one article URL per line (blank lines and lines
starting with # are ignored), treated as one unnamed pillar.

If given a directory: every *.txt file inside is treated as one pillar
(the filename, minus extension, is the pillar name), each holding its
own list of article URLs in the same format. This is the normal mode
for the recurring quarterly skim, e.g.:

    urls/biomedical-engineering.txt
    urls/journalism.txt
    urls/law-school.txt
    urls/premed.txt
"""

import re
import sys
import time
from datetime import date
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

HYPERLINK_FONT = Font(color="0563C1", underline="single")
TOP_ALIGN = Alignment(vertical="top", wrap_text=True)

# Column layout for the "Flagged stats" sheet. ARTICLE_COL holds one
# clickable hyperlink per row (one cell = one hyperlink, an Excel
# limitation); every other column is merged vertically across a stat's
# rows so a multi-article stat still reads as one grouped block.
SHEET_HEADER = ["Pillar", "SUMMARY", "Article", "Block Type", "Quote / Context",
                "Detected Year(s)", "Apparent Source", "Stat Type",
                "Suggested Cadence", "Staleness Assessment"]
ARTICLE_COL = 3
MERGE_COLS = [c for c in range(1, len(SHEET_HEADER) + 1) if c != ARTICLE_COL]

HEADERS = {"User-Agent": "Mozilla/5.0 (pillar-stats-extractor)"}

YEAR_RE = re.compile(r"\b(19|20)\d{2}\b")
NUMBER_RE = re.compile(
    r"\$\s?\d[\d,]*(\.\d+)?"       # dollar amounts
    r"|\d[\d,]*(\.\d+)?\s?%"       # percentages
    r"|\b\d{1,3}(,\d{3})+\b"       # comma-grouped numbers (enrollment, etc.)
)

SOURCE_PATTERNS = [
    ("IPEDS", re.compile(r"\bIPEDS\b", re.I)),
    ("BLS", re.compile(r"Bureau of Labor Statistics|\bBLS\b", re.I)),
    ("U.S. News & World Report", re.compile(r"U\.?S\.? News", re.I)),
    ("Niche", re.compile(r"\bNiche\b")),
    ("PayScale", re.compile(r"PayScale", re.I)),
    ("College Scorecard", re.compile(r"College Scorecard", re.I)),
]

STAT_TYPE_PATTERNS = [
    ("Wage", re.compile(r"salary|salaries|wage|pay\b|earn|income", re.I)),
    ("Outlook", re.compile(r"outlook|growth|demand|projected|job growth|decline", re.I)),
    # Checked before Enrollment: IPEDS methodology boilerplate almost always
    # says "...degree completions in X..." to describe how institutions were
    # selected/ranked, even when the actual displayed figures are costs --
    # without this ordering that boilerplate wins over the real "tuition"/
    # "cost" keyword every time.
    ("Cost/Tuition", re.compile(r"tuition|cost\b|price", re.I)),
    ("Enrollment", re.compile(r"enroll|completions?|graduates?", re.I)),
    ("Ranking", re.compile(r"rank(ed|ing)?|top \d", re.I)),
]

# Catches stat-bearing sentences even when no year is stated nearby, e.g.
# "median pay is $106,950" (no "2024" in the sentence).
STAT_KEYWORD_RE = re.compile(
    r"|".join(p.pattern for _, p in STAT_TYPE_PATTERNS)
    + r"|median|average|mean\b|according to",
    re.I,
)

# Lead-in sentences that introduce a table/stat rather than citing one
# themselves, e.g. "Below you'll find the top 10 colleges...". These get
# flagged as false positives if they happen to mention a threshold number
# ("20,000 total enrollment") near a stats keyword ("tuition").
INTRO_PHRASE_RE = re.compile(
    r"below you.ll find|you.ll find the|here are the|the following table"
    r"|the table below|as shown below|see the table",
    re.I,
)

NOISE_CLASS_HINTS = ("breadcrumb", "webform", "menu", "nav")
SOURCE_LINE_RE = re.compile(r"Source\s*:.*?[.!?]", re.I)

# Periods inside these should not be treated as sentence boundaries.
_PROTECTED_ABBREVIATIONS = [
    "U.S.A.", "U.S.", "U.K.", "Ph.D.", "M.D.", "vs.", "etc.",
    "approx.", "Dr.", "Mr.", "Ms.", "Mrs.", "Jr.", "Sr.", "St.",
    "Inc.", "Ltd.", "Co.",
]
_ABBREV_PLACEHOLDER = "․"  # one dot leader -- visually a period, not matched by [.!?]


def guess_source(text):
    for label, pattern in SOURCE_PATTERNS:
        if pattern.search(text):
            return label
    return "Unknown (needs review)"


def guess_stat_type(text):
    for label, pattern in STAT_TYPE_PATTERNS:
        if pattern.search(text):
            return label
    return "Other"


def suggested_cadence(source_label):
    if source_label in ("IPEDS", "BLS"):
        return "Annual"
    return "Review"


def latest_bls_wage_year(today=None):
    # BLS OEWS publishes new annual wage estimates each spring (historically
    # ~April), covering May-of-prior-year data. Heuristic, not a live check
    # of BLS's release calendar.
    today = today or date.today()
    return today.year - 1 if today.month >= 4 else today.year - 2


def latest_ipeds_completions_year(today=None):
    # IPEDS Completions data for academic year YYYY-(YY+1) is broadly
    # available roughly mid-the-following-year. Heuristic, not a live check
    # of IPEDS's release calendar. Year returned is the academic year's
    # *start* year (matches how these articles cite it, e.g. "2023-24").
    today = today or date.today()
    return today.year - 1 if today.month >= 7 else today.year - 2


def assess_staleness(source_label, years_str, today=None):
    if source_label == "Unknown (needs review)":
        return "Source unclear — manual check"
    if source_label not in ("BLS", "IPEDS"):
        # A real, named source (College Scorecard, Niche, PayScale, ...) --
        # just one we don't have a release-cadence heuristic for yet, which
        # is a different (and much better) situation than not knowing the
        # source at all.
        return f"{source_label} cited — no automated freshness check yet, verify manually"
    if not years_str:
        return "No year found — manual check"

    year_ints = [int(y) for y in years_str.split(", ") if y.strip().isdigit()]
    if not year_ints:
        return "No year found — manual check"

    cited_year = max(year_ints)
    latest_expected = (
        latest_bls_wage_year(today) if source_label == "BLS"
        else latest_ipeds_completions_year(today)
    )

    if cited_year < latest_expected:
        return (f"Likely stale — article cites {cited_year}, "
                f"~{latest_expected} data should now be available")
    if cited_year == latest_expected:
        return "Current — matches latest expected release"
    return "Current — cites data newer than heuristic expects (verify)"


def is_noise(el):
    ident = " ".join((el.get("class") or []) + [el.get("id") or ""]).lower()
    return any(hint in ident for hint in NOISE_CLASS_HINTS)


def protect_abbreviation_periods(text):
    protected = text
    for abbr in _PROTECTED_ABBREVIATIONS:
        protected = protected.replace(abbr, abbr.replace(".", _ABBREV_PLACEHOLDER))
    return protected


def split_sentences(text):
    parts = re.split(r"(?<=[.!?])\s+", protect_abbreviation_periods(text))
    return [p.replace(_ABBREV_PLACEHOLDER, ".") for p in parts]


def find_source_sentence(text):
    """Isolates just the "Source: ..." sentence out of a longer blob of
    text. Abbreviation periods (e.g. "U.S." in "Source: U.S. Bureau of
    Labor Statistics") are protected first -- SOURCE_LINE_RE's lazy match
    otherwise stops at the first period it finds, which without this would
    truncate that example down to just "Source: U.", losing the actual
    source name entirely."""
    match = SOURCE_LINE_RE.search(protect_abbreviation_periods(text))
    return match.group(0).replace(_ABBREV_PLACEHOLDER, ".") if match else None


US_STATES = [
    "Alabama", "Alaska", "Arizona", "Arkansas", "California", "Colorado",
    "Connecticut", "Delaware", "Florida", "Georgia", "Hawaii", "Idaho",
    "Illinois", "Indiana", "Iowa", "Kansas", "Kentucky", "Louisiana",
    "Maine", "Maryland", "Massachusetts", "Michigan", "Minnesota",
    "Mississippi", "Missouri", "Montana", "Nebraska", "Nevada",
    "New Hampshire", "New Jersey", "New Mexico", "New York",
    "North Carolina", "North Dakota", "Ohio", "Oklahoma", "Oregon",
    "Pennsylvania", "Rhode Island", "South Carolina", "South Dakota",
    "Tennessee", "Texas", "Utah", "Vermont", "Virginia", "Washington",
    "West Virginia", "Wisconsin", "Wyoming",
]


def describe_table(table_tag):
    """Pulls the table's <caption> and its <thead> column names, e.g.
    ("Salary ranges for biomedical engineers throughout the United
    States", ["Geography", "10th Percentile", "Median", "90th Percentile"])
    -- built from the actual table structure rather than the flattened
    " | "-joined text, since a header cell's own internal markup (e.g.
    "10th"<br/>"Percentile" split across nested <span> tags) would
    otherwise get mis-split into separate tokens by a plain text split."""
    caption_tag = table_tag.find("caption")
    caption = caption_tag.get_text(" ", strip=True) if caption_tag else None

    header_row = None
    thead = table_tag.find("thead")
    if thead:
        header_row = thead.find("tr")
    if header_row is None:
        header_row = table_tag.find("tr")

    headers = ([c.get_text(" ", strip=True) for c in header_row.find_all(["th", "td"])]
               if header_row else [])
    return caption, headers


def table_to_readable_text(table_tag):
    """Renders a table as one line per row (caption first, if any) instead
    of a single flattened " | "-joined wall of text -- with wrap_text on,
    Excel then shows it as an actual readable mini-table inside the cell."""
    caption, _ = describe_table(table_tag)
    lines = [caption] if caption else []
    for tr in table_tag.find_all("tr"):
        cells = [c.get_text(" ", strip=True) for c in tr.find_all(["th", "td"])]
        if cells:
            lines.append(" | ".join(cells))
    return "\n".join(lines)


def guess_table_name(table_tag):
    caption, headers = describe_table(table_tag)
    name = caption or (headers[0] if headers else "Data table")
    # First column is usually the row-label dimension (Geography,
    # Institution, ...), not a measured figure -- the remaining columns are
    # what you'd actually need to go look up, e.g. "10th Percentile, Median,
    # 90th Percentile".
    metrics = headers[1:] if len(headers) > 1 else []
    if metrics:
        name += ": " + ", ".join(metrics)
    return name


def shorten_name(name, max_len=90):
    """Hard length cap for the SUMMARY column -- a multi-caption stat
    widget or a wide table can otherwise join enough column names/captions
    to produce a genuinely long string. Cuts at the last comma/pipe/colon
    boundary within the limit so it reads as "X, Y, Z (+N more)" rather
    than an arbitrary mid-word chop."""
    if len(name) <= max_len:
        return name
    truncated = name[:max_len]
    for sep in (", ", " | ", ": "):
        idx = truncated.rfind(sep)
        if idx > max_len * 0.4:
            return truncated[:idx] + "..."
    idx = truncated.rfind(" ")
    return (truncated[:idx] if idx > 0 else truncated) + "..."


def guess_stat_name(row):
    """Best-effort short label for the stat -- a starting point to hand-edit,
    not a substitute for human review."""
    if row["block_type"] in ("Table", "StatWidget"):
        return shorten_name(row.get("_stat_name_hint") or "Statistic")

    quote, stat_type = row["quote"], row["stat_type"]
    geography = None
    if re.search(r"United States|\bU\.S\.", quote):
        geography = "US"
    else:
        for state in US_STATES:
            if re.search(rf"\b{re.escape(state)}\b", quote):
                geography = state
                break

    base = stat_type if stat_type != "Other" else "Stat"
    return f"{base} ({geography})" if geography else base


def find_main_content(soup):
    article = soup.find("article")
    if article:
        return article
    return soup.find("main") or soup.find("body")


def find_trailing_context(start_el, max_hops=3, max_climbs=3):
    """Walks forward through siblings -- climbing out of wrapper divs
    (e.g. div.table-responsive, a layout column with no sibling of its
    own) when a level runs out of siblings -- collecting text until a
    "Source: ..." line is found or the hop/climb budget runs out."""
    context = ""
    node = start_el
    hops = 0
    climbs = 0
    while hops < max_hops:
        nxt = node.find_next_sibling()
        if nxt is None:
            if climbs < max_climbs and node.parent is not None and node.parent.name in ("div", "section"):
                node = node.parent
                climbs += 1
                continue
            break
        nxt_text = nxt.get_text(" ", strip=True)
        if nxt_text:
            context += " || " + nxt_text
            if re.search(r"source\s*:", nxt_text, re.I):
                break
        node = nxt
        hops += 1
    return context


def extract_stat_widgets(content):
    """Some pages use a "statistic callout" component instead of a table:
    a div.stat__grid per number, with the figure in h2.stat__title (often
    split across a "$" prefix span and a digits span) and its caption in a
    sibling p.stat__content, e.g.:

        <div class="stat stat__grid">
          <div><h2 class="stat__title">
            <span class="headline__prefix">$</span>
            <span class="headline__heading">13,131</span>
          </h2></div>
          <p class="stat__content">Public, in-state college/university</p>
        </div>

    Multiple such boxes sitting in the same layout row (e.g. a 3-column
    row) are grouped into one flagged entry, same as a multi-column table."""
    groups = {}
    order = []
    for grid in content.find_all("div", class_="stat__grid"):
        h2 = grid.find("h2", class_="stat__title")
        if h2 is None:
            continue

        # Each box in a multi-column row (e.g. 3 side-by-side columns) can
        # be wrapped in its own per-column layout div at a different depth,
        # so matching on a "layout--*" ancestor class picks a different,
        # unshared wrapper per box. The nearest preceding non-widget H2 is
        # the section heading introducing this row -- a far more reliable
        # signal that boxes belong together than DOM-ancestor guessing.
        # (Not using class_=<lambda> here: BS4 calls a class_ function once
        # per individual class token, not with the tag's full class list --
        # so a negative check like "stat__title not in classes" can match
        # on an unrelated second class token and pick up another stat box.)
        heading_tag = next(
            (h for h in grid.find_all_previous("h2")
             if "stat__title" not in (h.get("class") or [])),
            None,
        )
        key = id(heading_tag) if heading_tag is not None else id(grid)
        if key not in groups:
            groups[key] = {"heading": heading_tag, "boxes": []}
            order.append(key)
        groups[key]["anchor"] = grid  # keep advancing to the last box seen in this group

        number = h2.get_text(strip=True)  # no separator -- avoids inserting
                                           # a space between "$" and digits
                                           # that live in separate spans
        # Caption can be split across an optional "description" span (e.g.
        # "Estimated growth rate") and the "content" paragraph (e.g. "for
        # biomedical engineering jobs from 2024 to 2034...") -- grab both.
        description_tag = grid.find("span", class_="stat__description")
        content_tag = grid.find("p", class_="stat__content")
        caption = " ".join(
            t.get_text(" ", strip=True) for t in (description_tag, content_tag) if t
        )
        groups[key]["boxes"].append((number, caption))

    rows = []
    for key in order:
        boxes = groups[key]["boxes"]
        if not boxes:
            continue
        anchor = groups[key]["anchor"]
        quote = "\n".join(f"{cap}: {num}" if cap else num for num, cap in boxes)

        # The trailing "Source: ..." line's own optional methodology
        # sentence isn't reliably present (some pages include it, some
        # don't), which made years/classification unstable across
        # duplicate copies of "the same" widget on different articles.
        # Anchor on the more stable signals instead: the short Source
        # sentence alone for the year, and the section heading that
        # actually introduces this widget (e.g. "How much does it cost to
        # study biomedical engineering in the US?") for classification.
        trailing = find_trailing_context(anchor)
        source_line = find_source_sentence(trailing) or trailing
        years = sorted(set(m.group(0) for m in YEAR_RE.finditer(f"{quote} {source_line}")))

        heading_tag = groups[key]["heading"]
        heading = heading_tag.get_text(" ", strip=True) if heading_tag else ""
        classify_text = f"{quote} {heading} {source_line}"

        stat_name_hint = " | ".join(cap for _, cap in boxes if cap) or "Statistic"

        rows.append({
            "block_type": "StatWidget",
            "quote": quote,
            "years": ", ".join(years),
            "source": guess_source(classify_text),
            "stat_type": guess_stat_type(classify_text),
            "_merge_key": None,
            "_stat_name_hint": stat_name_hint,
        })
    return rows


def fetch_article(url):
    resp = requests.get(url, headers=HEADERS, timeout=20)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "lxml")

    content = find_main_content(soup)
    if content is None:
        return url, []

    title_tag = content.find("h1") or soup.title
    title = title_tag.get_text(strip=True) if title_tag else url

    for tag in content.find_all(["script", "style", "form", "nav"]):
        tag.decompose()

    rows = []
    seen_tables = set()

    blocks = content.find_all(["p", "li", "h2", "h3", "h4", "table"])
    for block in blocks:
        if is_noise(block):
            continue

        if block.name == "table":
            if id(block) in seen_tables:
                continue
            seen_tables.add(id(block))

            table_text = table_to_readable_text(block)
            context_text = table_text + find_trailing_context(block)

            years = sorted(set(m.group(0) for m in YEAR_RE.finditer(context_text)))
            if not years and not STAT_KEYWORD_RE.search(context_text) and not NUMBER_RE.search(context_text):
                continue

            quote = table_text

            tail = context_text[len(table_text):]
            source_sentence = find_source_sentence(tail)
            merge_key = source_sentence.strip().lower() if source_sentence else (tail or None)

            rows.append({
                "block_type": "Table",
                "quote": quote,
                "years": ", ".join(years),
                "source": guess_source(context_text),
                "stat_type": guess_stat_type(context_text),
                "_merge_key": merge_key,
                "_stat_name_hint": guess_table_name(block),
            })
            continue

        text = block.get_text(" ", strip=True)
        if not text:
            continue

        for sentence in split_sentences(text):
            if INTRO_PHRASE_RE.search(sentence):
                continue
            has_number = NUMBER_RE.search(sentence)
            has_year = YEAR_RE.search(sentence)
            if has_number and (has_year or STAT_KEYWORD_RE.search(sentence)):
                years = sorted(set(m.group(0) for m in YEAR_RE.finditer(sentence)))
                quote = sentence if len(sentence) <= 300 else sentence[:300] + "..."
                rows.append({
                    "block_type": "Paragraph",
                    "quote": quote,
                    "years": ", ".join(years),
                    "source": guess_source(sentence),
                    "stat_type": guess_stat_type(sentence),
                    "_merge_key": None,
                })

    rows.extend(extract_stat_widgets(content))
    return title, merge_sibling_tables(rows)


def merge_sibling_tables(rows):
    """Adjacent Table rows that share the same trailing source/caption text
    (e.g. two tables both captioned "Source: IPEDS, 2023-24 completions...")
    are presenting one dataset split across tables for display -- collapse
    them into a single row."""
    merged = []
    for row in rows:
        key = row.get("_merge_key")
        if (key and merged and merged[-1].get("block_type") == "Table"
                and merged[-1].get("_merge_key") == key):
            merged[-1]["quote"] += "\n\n" + row["quote"]
        else:
            merged.append(row)
    for row in merged:
        row.pop("_merge_key", None)
    return merged


def load_urls(path):
    urls = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#"):
                urls.append(line)
    return urls


def pillar_name_from_filename(path):
    return path.stem.replace("-", " ").replace("_", " ").title()


def gather_pillar_url_lists(input_path):
    """Returns list of (pillar_name, [urls])."""
    path = Path(input_path)
    if path.is_dir():
        pillars = []
        for txt_file in sorted(path.glob("*.txt")):
            pillars.append((pillar_name_from_filename(txt_file), load_urls(txt_file)))
        return pillars
    return [(pillar_name_from_filename(path), load_urls(path))]


def default_output_path():
    return f"Pillar Stat Review - {date.today():%b %d %Y}.xlsx"


def stat_fingerprint(row):
    """Identifies "the same stat" by its actual numbers/year rather than
    exact quote text (which often gets paraphrased slightly between
    articles -- "the U.S." vs "the United States", "between 2024 and 2034"
    vs "from 2024 to 2034") or by source/stat_type (a page that reuses a
    widget/table sometimes drops its "Source: ..." citation entirely, which
    would otherwise fingerprint the same real stat differently depending on
    which article happened to still include it)."""
    numbers = tuple(sorted(m.group(0).strip() for m in NUMBER_RE.finditer(row["quote"])))
    return (row["block_type"], row["years"], numbers)


def dedupe_across_articles(pillar_rows):
    """pillar_rows: list of (row_dict, title, url). Collapses rows that are
    the same underlying stat (same fingerprint) cited in multiple articles
    into one row with all locations combined. If one occurrence has a real
    source/stat_type and another (of the same underlying stat) doesn't,
    the group upgrades to the more informative one."""
    by_key = {}
    order = []
    for row, title, url in pillar_rows:
        key = stat_fingerprint(row)
        if key not in by_key:
            by_key[key] = {**row, "locations": []}
            order.append(key)
        elif (by_key[key]["source"] == "Unknown (needs review)"
              and row["source"] != "Unknown (needs review)"):
            by_key[key]["source"] = row["source"]
            by_key[key]["stat_type"] = row["stat_type"]
        by_key[key]["locations"].append((title, url))
    return [by_key[key] for key in order]


def write_stat_group(ws, pillar_name, row):
    """Writes one deduped stat as N rows (one per article it appears in,
    each with a real clickable hyperlink), then merges every other column
    vertically across those rows so the group still reads as one entry.
    Returns True if this stat is likely stale."""
    stat_name = row["stat_name"]
    staleness = assess_staleness(row["source"], row["years"])
    shared = [pillar_name, stat_name, row["block_type"], row["quote"],
              row["years"], row["source"], row["stat_type"],
              suggested_cadence(row["source"]), staleness]

    start_row = ws.max_row + 1
    for title, url in row["locations"]:
        ws.append(shared[:2] + [title] + shared[2:])
        r = ws.max_row
        link_cell = ws.cell(row=r, column=ARTICLE_COL)
        link_cell.hyperlink = url
        link_cell.font = HYPERLINK_FONT
        for col in range(1, len(SHEET_HEADER) + 1):
            ws.cell(row=r, column=col).alignment = TOP_ALIGN

    end_row = ws.max_row
    if end_row > start_row:
        for col in MERGE_COLS:
            ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

    return staleness.startswith("Likely stale")


def main():
    if len(sys.argv) not in (2, 3):
        print("Usage: python extract_stats.py <urls.txt | urls_dir> [output.xlsx]")
        print(f"  (output defaults to '{default_output_path()}' if omitted)")
        sys.exit(1)

    input_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) == 3 else default_output_path()
    pillar_url_lists = gather_pillar_url_lists(input_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Flagged stats"
    ws.append(SHEET_HEADER)
    for col in range(1, len(SHEET_HEADER) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    total_flagged = 0
    total_stale = 0
    for pillar_name, urls in pillar_url_lists:
        print(f"=== Pillar: {pillar_name} ({len(urls)} article(s)) ===")
        pillar_rows = []
        for i, url in enumerate(urls, 1):
            print(f"[{i}/{len(urls)}] Fetching {url}")
            try:
                title, rows = fetch_article(url)
            except Exception as exc:
                print(f"  ERROR: {exc}")
                ws.append([pillar_name, "", url, "ERROR", str(exc), "", "", "", "", ""])
                ws.cell(row=ws.max_row, column=ARTICLE_COL).hyperlink = url
                ws.cell(row=ws.max_row, column=ARTICLE_COL).font = HYPERLINK_FONT
                continue

            print(f"  {len(rows)} flagged stat(s)")
            for row in rows:
                pillar_rows.append((row, title, url))

            time.sleep(0.5)  # be polite to the server

        deduped = dedupe_across_articles(pillar_rows)
        for row in deduped:
            row["stat_name"] = guess_stat_name(row)
        # Sort (stable) so rows sharing a stat name are always contiguous --
        # e.g. two distinct "Wage (US)" facts should sit next to each other
        # rather than being separated by an unrelated table row, so Lydia
        # can eyeball and manually merge them together if she wants to.
        deduped.sort(key=lambda r: r["stat_name"].lower())

        total_flagged += len(deduped)
        for row in deduped:
            if write_stat_group(ws, pillar_name, row):
                total_stale += 1

    widths = [20, 32, 40, 10, 55, 14, 22, 14, 16, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wb.save(out_path)
    total_articles = sum(len(urls) for _, urls in pillar_url_lists)
    print(f"\nDone. {total_flagged} unique stat(s) flagged across {total_articles} article(s) "
          f"in {len(pillar_url_lists)} pillar(s).")
    print(f"{total_stale} flagged as likely stale.")
    print(f"Saved to {out_path}")


if __name__ == "__main__":
    main()
