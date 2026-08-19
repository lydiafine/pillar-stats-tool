"""
Fetches College Scorecard field-of-study earnings/debt data for a given CIP
code, aggregated across every institution offering it -- replicating the
documented methodology (see reference/BME Statistics for Pillar Articles.docx):

    "Using College Scorecard data, we pulled all institutions with
    completions in bachelor's or master's degrees, then averaged the 1-year
    and 5-year post-grad earnings reported by the institution, as well as
    median federal student debt, for both bachelor's and master's."

Field mapping verified 2026-08-19 against College Scorecard's own
FieldOfStudy_Data_Dictionary (collegescorecard.ed.gov/files/
CollegeScorecardDataDictionary.xlsx):
    CIPCODE          -> cip_4_digit.code
    CREDLEV          -> cip_4_digit.credential.level (3=Bachelor's, 5=Master's
                         -- confirmed via live API response labels, not just
                         the dictionary's truncated value/label table)
    EARN_MDN_1YR     -> cip_4_digit.earnings.1_yr.overall_median_earnings
    EARN_MDN_5YR     -> cip_4_digit.earnings.5_yr.overall_median_earnings
    DEBT_ALL_STGP_ANY_MDN -> cip_4_digit.debt.staff_grad_plus.all.all_inst.median
                         (median Stafford + Grad PLUS loan debt, source NSLDS)

This is a genuine cross-institution average, not a single lookup -- expect
hundreds of institutions per CIP code, most with partial (null) data for any
given metric. Averages are computed only over institutions that actually
report that specific metric, matching "averaged ... reported by the
institution."

Usage:
    python fetch_college_scorecard.py <CIP-4-digit> [output.xlsx] [--baseline=path.json]
    python fetch_college_scorecard.py 1405 "CS Earnings Data.xlsx" --baseline=reference/bme_scorecard_baseline.json

Requires a College Scorecard API key (free, instant at api.data.gov/signup).
Set the CS_API_KEY environment variable.
"""

import json
import os
import statistics
import sys

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

API_BASE = "https://api.data.gov/ed/collegescorecard/v1/schools.json"
PER_PAGE = 100

CREDENTIAL_LEVELS = {
    3: "Bachelor's",
    5: "Master's",
}

FIELDS = ",".join([
    "school.name",
    "latest.programs.cip_4_digit.credential.level",
    "latest.programs.cip_4_digit.earnings.1_yr.overall_median_earnings",
    "latest.programs.cip_4_digit.earnings.5_yr.overall_median_earnings",
    "latest.programs.cip_4_digit.debt.staff_grad_plus.all.all_inst.median",
])

REVIEW_THRESHOLD_PCT = 10


def fetch_all_programs(cip_code, api_key):
    """Paginates through every institution offering this CIP code, returning
    one dict per (institution, credential level) program record."""
    programs = []
    page = 0
    while True:
        resp = requests.get(API_BASE, params={
            "latest.programs.cip_4_digit.code": cip_code,
            "fields": FIELDS,
            "per_page": PER_PAGE,
            "page": page,
            "api_key": api_key,
        }, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        for school in data["results"]:
            school_name = school.get("school.name")
            for prog in school.get("latest.programs.cip_4_digit", []):
                prog["_school"] = school_name
                programs.append(prog)
        total = data["metadata"]["total"]
        if (page + 1) * PER_PAGE >= total:
            break
        page += 1
    return programs


def safe_get(prog, *path):
    node = prog
    for key in path:
        if not isinstance(node, dict):
            return None
        node = node.get(key)
    return node


def aggregate(programs):
    """Returns {level: {metric: (mean, n)}} for each credential level of
    interest, averaging only over institutions with a non-null value for
    that specific metric."""
    by_level = {level: [] for level in CREDENTIAL_LEVELS}
    for prog in programs:
        level = safe_get(prog, "credential", "level")
        if level in by_level:
            by_level[level].append(prog)

    metrics = {
        "earn_1yr": ("earnings", "1_yr", "overall_median_earnings"),
        "earn_5yr": ("earnings", "5_yr", "overall_median_earnings"),
        "debt": ("debt", "staff_grad_plus", "all", "all_inst", "median"),
    }

    results = {}
    for level, progs in by_level.items():
        results[level] = {"institution_count": len(progs)}
        for metric_name, path in metrics.items():
            values = [v for v in (safe_get(p, *path) for p in progs) if v is not None]
            results[level][metric_name] = (
                (round(statistics.mean(values)), len(values)) if values else (None, 0)
            )
    return results


def load_baseline(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)["values"]


def pct_change(new, old):
    if old in (None, 0):
        return None
    return round((new - old) / old * 100, 1)


def save_to_xlsx(results, cip_code, out_path, baseline=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "College Scorecard"

    header = ["Credential Level", "Institutions Reporting", "1yr Earnings",
               "5yr Earnings", "Median Debt (Stafford+GradPLUS)"]
    has_baseline = baseline is not None
    if has_baseline:
        header += ["Baseline 1yr", "% Chg 1yr", "Baseline 5yr", "% Chg 5yr", "Review"]

    ws.append([f"CIP {cip_code}"])
    ws.append(header)
    for col in range(1, len(header) + 1):
        ws.cell(row=2, column=col).font = Font(bold=True)

    for level, label in CREDENTIAL_LEVELS.items():
        r = results[level]
        earn1, n1 = r["earn_1yr"]
        earn5, n5 = r["earn_5yr"]
        debt, ndebt = r["debt"]
        row = [
            label,
            r["institution_count"],
            f"{earn1:,} (n={n1})" if earn1 is not None else "no data",
            f"{earn5:,} (n={n5})" if earn5 is not None else "no data",
            f"{debt:,} (n={ndebt})" if debt is not None else "no data",
        ]
        if has_baseline:
            base = baseline.get(label, {})
            b1, b5 = base.get("earn_1yr"), base.get("earn_5yr")
            c1, c5 = pct_change(earn1, b1), pct_change(earn5, b5)
            review = []
            if c1 is not None and abs(c1) > REVIEW_THRESHOLD_PCT:
                review.append(f"1yr earnings {c1:+.1f}%")
            if c5 is not None and abs(c5) > REVIEW_THRESHOLD_PCT:
                review.append(f"5yr earnings {c5:+.1f}%")
            row += [b1, c1, b5, c5, "; ".join(review) if review else ""]
        ws.append(row)

    widths = [16, 20, 18, 18, 28]
    if has_baseline:
        widths += [14, 12, 14, 12, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)


def main():
    baseline_path = None
    positional = []
    for arg in sys.argv[1:]:
        if arg.startswith("--baseline="):
            baseline_path = arg.split("=", 1)[1]
        else:
            positional.append(arg)

    if len(positional) < 1:
        print("Usage: python fetch_college_scorecard.py <CIP-4-digit> [output.xlsx] [--baseline=path.json]")
        sys.exit(1)

    cip_code = positional[0]
    out_path = positional[1] if len(positional) > 1 else f"College Scorecard Data - CIP {cip_code}.xlsx"

    api_key = os.environ.get("CS_API_KEY")
    if not api_key:
        print("Set the CS_API_KEY environment variable (free key from api.data.gov/signup).")
        sys.exit(1)

    print(f"Fetching all institutions with CIP {cip_code} programs...")
    programs = fetch_all_programs(cip_code, api_key)
    print(f"Fetched {len(programs)} program records.")

    results = aggregate(programs)
    baseline = load_baseline(baseline_path) if baseline_path else None
    save_to_xlsx(results, cip_code, out_path, baseline)

    print(f"\nDone. Saved to {out_path}")
    for level, label in CREDENTIAL_LEVELS.items():
        r = results[level]
        earn1, n1 = r["earn_1yr"]
        earn5, n5 = r["earn_5yr"]
        debt, ndebt = r["debt"]
        print(f"  {label} ({r['institution_count']} institutions):")
        print(f"    1yr earnings: {'$'+format(earn1, ',') if earn1 else '?'} (n={n1})")
        print(f"    5yr earnings: {'$'+format(earn5, ',') if earn5 else '?'} (n={n5})")
        print(f"    median debt:  {'$'+format(debt, ',') if debt else '?'} (n={ndebt})")


if __name__ == "__main__":
    main()
